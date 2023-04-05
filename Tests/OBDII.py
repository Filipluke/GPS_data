import obd
import time
import matplotlib.pyplot as plt
import openpyxl

# Konfiguracja połączenia z portem OBD II

connection = obd.OBD() 
#excel

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="lp")
ws.cell(row=1, column=2, value="V[Km/h]")
ws.cell(row=1, column=3, value="t[s]")

# Konfiguracja wykresu

plt.ion()
fig, ax = plt.subplots()
plt.title('Wykres prędkości')
plt.xlabel('Czas [s]')
plt.ylabel('Prędkość [km/h]')
plt.xlim(0, 60) # ustawienie granic osi X
plt.ylim(0, 200) # ustawienie granic osi Y
line, = ax.plot([], [], lw=2)

# Odczytanie wartości prędkości co 1 sekundę i rysowanie na wykresie w czasie rzeczywistym
xdata, ydata = [], []
i=0
time=0
while True:
    try:
        speed = connection.query(obd.commands.SPEED).value.to('km/h')
        xdata.append(time.time() - xdata[0] if xdata else 0)
        ydata.append(speed.magnitude)
        line.set_data(xdata, ydata)
        plt.draw()
        plt.pause(0.001)

        ws.cell(row=i+2, column=1, value=i+1)
        ws.cell(row=i+2, column=2, value=speed)
        ws.cell(row=i+2, column=3, value=time+0.1)




    except:
        print("Błąd odczytu wartości prędkości.")
        wb.save('Pomiary_z_OBD.xlsx')
        break