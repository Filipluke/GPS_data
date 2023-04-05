from pathlib import Path
import serial.tools.list_ports
from tkinter import *
import math
from typing import List, Tuple
import numpy as np
import matplotlib.pyplot as plt
import threading
import tkinter.font as tkFont
from tkinter import ttk
import openpyxl
import os
import time

BASE_PATH = Path(os.path.dirname(os.path.abspath(__file__)))
ASSETS_PATH = BASE_PATH / "assets" / "frame0"

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()

window.geometry("478x698")
window.configure(bg = "#FFFFFF")

#### 
global V
global Czas
V = []
Czas = []
keepRunning= True
####

def Export():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="lp")
    ws.cell(row=1, column=2, value="V[Km/h]")
    ws.cell(row=1, column=3, value="t[s]")
    for i in range(len(V)-1):
        ws.cell(row=i+2, column=1, value=i+1)
        ws.cell(row=i+2, column=2, value=V[i])
        ws.cell(row=i+2, column=3, value=Czas[i])
    wb.save('Velocity.xlsx')

def Start():
    global keepRunning
    global Czas
    serialInst = serial.Serial()
    serialInst.baudrate = 9600
    serialInst.port = ("COM"+entry_1.get())
    serialInst.open()
    

    fig, ax = plt.subplots()
    line, = ax.plot([], [])
    ax.set_xlim(0, 10) 
    ax.set_ylim(0, 100)
    ax.set_xlabel('Czas [s]')  
    ax.set_ylabel('Prędkość [Km/h]')
    ax.set_title('Prędkość w czasie rzeczywistym')
    xdata, ydata = [], []
    plt.show(block=False)

    start_time = time.time()
    
    while keepRunning:
        try:
            if serialInst.in_waiting:
                packet = serialInst.readline()
                V_decoded = float(packet.decode('utf').rstrip('\n'))
                print(V_decoded)


                y = float(V_decoded)
                x = time.time() - start_time
                xdata.append(x)
                ydata.append(y)
                line.set_data(xdata, ydata)
                ax.relim()
                ax.autoscale_view(True,True,True)
                ax.set_xlim(x-10, x)  # aktualizacja osi X
                fig.canvas.draw()
                fig.canvas.flush_events()

                if V_decoded != 0.0:
                    V.append(V_decoded)
                    Czas.append(x)
                    
                entry_3.delete(0, END)
                entry_3.insert(0, round(V_decoded, 1))
                
                
                
               
        except serial.SerialException:  
            print("Urządzenie zostało odłączone")
            keepRunning = False


def Break():
    global keepRunning
    keepRunning = False
    
 

def on_button1_clicked_Start():
    global keepRunning
    keepRunning = True
    threading.Thread(target=Start).start()



canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 698,
    width = 478,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    0.0,
    478.0,
    698.0,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    554.0,
    215.0,
    anchor="nw",
    text="Filip Żelaźnicki 2023",
    fill="#FFFFFF",
    font=("Inter Bold", 22 * -1)
)

canvas.create_text(
    140.0,
    45.0,
    anchor="nw",
    text="Wpisz dane",
    fill="#000000",
    font=("Inter Bold", 36 * -1)
)

entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    225.5,
    148.5,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0,
    justify="center",
    font=("Inter Regular", 20 * -1)
)
entry_1.place(
    x=186.0,
    y=128.0,
    width=79.0,
    height=39.0
)




entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    242.5,
    336.5,
    image=entry_image_3
)
entry_3 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#FF1300",
    highlightthickness=0,
    justify="center",
    font=("Inter Regular", 100 * -1)
)
entry_3.place(
    x=126.0,
    y=250.0,
    width=233.0,
    height=171.0
)



canvas.create_text(
    180.0,
    100.0,
    anchor="nw",
    text="Numer portu",
    fill="#000000",
    font=("Inter Regular", 16 * -1)
)

canvas.create_text(
    7.0,
    7.0,
    anchor="nw",
    text="Filip Żelaźnicki 2023",
    fill="#020202",
    font=("Inter Regular", 16 * -1)
)

canvas.create_text(
    153.0,
    214.0,
    anchor="nw",
    text="Prędkość [Km/h]",
    fill="#000000",
    font=("Inter Regular", 20 * -1)
)

button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: on_button1_clicked_Start(),
    relief="flat"
)
button_1.place(
    x=16.0,
    y=493.0,
    width=209.0,
    height=62.0
)

button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: Break(),
    relief="flat"
)
button_2.place(
    x=248.0,
    y=493.0,
    width=209.0,
    height=62.0
)

button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
   command=lambda: Export(),
    relief="flat"
)
button_3.place(
    x=103.0,
    y=594.0,
    width=272.0,
    height=74.09524536132812
)
window.resizable(False, False)
window.mainloop()
