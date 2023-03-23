from pathlib import Path
import serial.tools.list_ports
from tkinter import *
import math
from typing import List, Tuple
import numpy as np
import matplotlib.pyplot as plt
import threading
import tkinter.font as tkFont
from tkinter import ttk
import openpyxl
import os


BASE_PATH = Path(os.path.dirname(os.path.abspath(__file__)))
ASSETS_PATH = BASE_PATH / "assets" / "frame0"



def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()


def getMps(vals: List[float]) -> List[float]:
    
    Vmps = [vals[i]/3.6 for i in range(len(vals))]
    return Vmps

def get_delta(vals: List[float]) -> List[float]:
    #frequency in hertz
    f=float(entry_8.get())
    dvals = [(vals[i+1] - vals[i])*f for i in range(len(vals)-1)]
    return dvals

def get_negatives(vals: List[float]) -> List[float]:
    vals_neg = [val for val in vals if val < 0]
    return vals_neg

def get_dkg(d_v: List[float]) -> List[float]:
    dkg = [0.1 * val for val in d_v]
    return dkg

def get_v2(v: List[float], dvNeg: List[float]) -> List[float]:
    v2 = [v[i+1] * v[i+1] for i in range(len(dvNeg))]
    return v2

def linear_regression(x_vals: List[float], y_vals: List[float]) -> Tuple[float, float, float]:
    if len(x_vals) != len(y_vals):
        raise ValueError("Input values differ in length.")
    
    sum_of_x = sum(x_vals)
    sum_of_y = sum(y_vals)
    sum_of_x_sq = sum(x*x for x in x_vals)
    sum_of_y_sq = sum(y*y for y in y_vals)
    sum_codeviates = sum(x*y for x,y in zip(x_vals, y_vals))

    count = len(x_vals)
    ss_x = sum_of_x_sq - ((sum_of_x * sum_of_x) / count)

    r_numerator = (count * sum_codeviates) - (sum_of_x * sum_of_y)
    r_denom = (count * sum_of_x_sq - (sum_of_x * sum_of_x)) * (count * sum_of_y_sq - (sum_of_y * sum_of_y))
    s_co = sum_codeviates - ((sum_of_x * sum_of_y) / count)

    mean_x = sum_of_x / count
    mean_y = sum_of_y / count
    dbl_r = r_numerator / math.sqrt(r_denom)

    r_squared = dbl_r * dbl_r
    y_intercept = mean_y - ((s_co / ss_x) * mean_x)
    slope = s_co / ss_x

    return r_squared, y_intercept, slope



#### 
global V
global time
V = []
keepRunning= True
####


def Export():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="lp")
    ws.cell(row=1, column=2, value="V[Km/h]")
    ws.cell(row=1, column=3, value="t[s]")
    for i in range(len(V)):
        ws.cell(row=i+2, column=1, value=i+1)
        ws.cell(row=i+2, column=2, value=V[i])
        ws.cell(row=i+2, column=3, value=time[i])
    wb.save('Pomiary.xlsx')
    
def Start():
    global keepRunning
    serialInst = serial.Serial()
    serialInst.baudrate = 9600
    serialInst.port = ("COM"+entry_6.get())
    serialInst.open()
    while keepRunning:
        try:
            if serialInst.in_waiting:
                packet = serialInst.readline()
                V_decoded = float(packet.decode('utf').rstrip('\n'))
                print(V_decoded)
                V.append(V_decoded)
                entry_7.delete(0, END)
                entry_7.insert(0, round(V_decoded, 1))
                
                
                
               
        except serial.SerialException:  
            print("Urządzenie zostało odłączone")
            keepRunning = False
def Break():
    global keepRunning
    keepRunning = False
    global time
    #Wyświetlenie Wykresu prędkości
    period= float(1/(float(entry_8.get())))
    time = [period * i for i in range(len(V))]
    plt.plot(time, V)
    plt.xlabel('Czas [s]')
    plt.ylabel('Prędkość[Km/h]')
    plt.show()
def on_button2_clicked_Start():
    global keepRunning
    keepRunning = True
    threading.Thread(target=Start).start()


def calculateCx():
  
    # data processing
    
    VMps = getMps(V)
    dv = get_delta(VMps)
    dvNeg = get_negatives(dv)
    dkg = get_dkg(dvNeg)
    v2 = get_v2(VMps, dvNeg)

    r2, b, a = linear_regression(v2, dkg)

    # constants
    m =float( entry_1.get())
    g = 9.81
    p = float( entry_2.get())
    A = float( entry_3.get())

    # m = 1200
    # p = 1.2
    # A = 2.2

    Cx = - (2*m*g*a) / (p*A)

    # data output
    print(f"Cx = {Cx}")
    print(V[0])
    print(VMps[0])
    

    
    approx = [a*x +b for x in v2]

    #approx_teoret = [-4.2e-05*x -0.0137 for x in v2]
    plt.plot(v2, approx)
    #plt.plot(v2, approx_teoret)
    plt.plot(v2, dkg, 'o')

    plt.xlabel("Kwadrat prędkości [m^2/s^2]")
    plt.ylabel("dk/g*a")
    
    plt.show()
    
    equation = str(np.polyfit(v2, dkg, 1))
    equation = equation.replace(" ", "x")
    equation = equation.replace("e", "*10^")
 

    print(equation)
    # v(t) [km/h]
    #plt.plot([x/1000 for x in t_8], [x for x in v_8])
    # a(t) [m/s^2]
    #plt.plot([x/1000 for x in t_8[1:]], [x/3.6 for x in dv_8])

    entry_4.insert(0, Cx)  # ustaw wartość zmiennej cx w polu tekstowym
    entry_5.insert(0, equation)  # ustaw wartość zmiennej cx w polu tekstowym


    


window.geometry("700x500")
window.configure(bg = "#FFFFFF")

canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 500,
    width = 700,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    0.0,
    241.0,
    500.0,
    fill="#2E8CE3",
    outline="")

canvas.create_text(
    11.0,
    458.0,
    anchor="nw",
    text="Filip Żelaźnicki 2023",
    fill="#FFFFFF",
    font=("Inter Bold", 22 * -1)
)

canvas.create_text(
    350.0,
    19.0,
    anchor="nw",
    text="Wpisz Dane",
    fill="#000000",
    font=("Inter Bold", 32 * -1)
)
canvas.create_text(
    21.0,
    131.0,
    anchor="nw",
    text="Prędkość Km/h",
    fill="#000000",
    font=("Inter Regular", 20 * -1)
)


entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    368.0,
    114.5,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_1.place(
    x=272.0,
    y=92.0,
    width=192.0,
    height=43.0
)

entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    369.0,
    184.5,
    image=entry_image_2
)
entry_2 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_2.place(
    x=273.0,
    y=162.0,
    width=192.0,
    height=43.0
)

entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    369.0,
    264.5,
    image=entry_image_3
)
entry_3 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0
)
entry_3.place(
    x=273.0,
    y=242.0,
    width=192.0,
    height=43.0
)

entry_image_4 = PhotoImage(
    file=relative_to_assets("entry_4.png"))
entry_bg_4 = canvas.create_image(
    599.5,
    186.0,
    image=entry_image_4
)
entry_4 = Entry(
    bd=0,
    bg="#D475D6",
    fg="#000716",
    highlightthickness=0
)
entry_4.place(
    x=519.0,
    y=162.0,
    width=161.0,
    height=46.0
)

entry_image_5 = PhotoImage(
    file=relative_to_assets("entry_5.png"))
entry_bg_5 = canvas.create_image(
    599.5,
    264.0,
    image=entry_image_5
)
entry_5 = Entry(
    bd=0,
    bg="#D575D7",
    fg="#000716",
    highlightthickness=0
)
entry_5.place(
    x=519.0,
    y=240.0,
    width=161.0,
    height=46.0
)
entry_image_6 = PhotoImage(
    file=relative_to_assets("entry_7.png"))
entry_bg_6 = canvas.create_image(
    117.5,
    92.5,
    image=entry_image_6
)
entry_6 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    highlightthickness=0,
    justify="center",
    font=("Inter Regular", 20 * -1)
    
)
entry_6.place(
    x=100.0,
    y=72.0,
    width=35.0,
    height=39.0
)

entry_image_7 = PhotoImage(
    file=relative_to_assets("entry_6.png"))
entry_bg_7 = canvas.create_image(
    121.5,
    237.5,
    image=entry_image_7
)
entry_7 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#FF1300",
    highlightthickness=0,
    justify="center",
    font=("Inter Regular", 72 * -1)
)
entry_7.place(
    x=25.0,
    y=164.0,
    width=193.0,
    height=145.0
)


entry_image_8 = PhotoImage(
    file=relative_to_assets("entry_8.png"))
entry_bg_8 = canvas.create_image(
    109.5,
    413.5,
    image=entry_image_8
)
entry_8 = Entry(
    bd=0,
    bg="#D9D9D9",
    fg="#000716",
    justify="center",
    font=("Inter Regular", 20 * -1),
    highlightthickness=0
)
entry_8.place(
    x=92.0,
    y=393.0,
    width=35.0,
    height=39.0
)

canvas.create_text(
    268.0,
    73.0,
    anchor="nw",
    text="Masa Samochodu [Kg]\n",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)

canvas.create_text(
    262.0,
    143.0,
    anchor="nw",
    text="Powierzchnia czołowa [m^2]",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)
canvas.create_text(
    22.0,
    30.0,
    anchor="nw",
    text="Port na którym znajduje się GPS",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)

canvas.create_text(
    268.0,
    221.0,
    anchor="nw",
    text="Gęstość powietrza [Kg/m^3]",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)

canvas.create_text(
    518.0,
    146.0,
    anchor="nw",
    text="Obliczony CX",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)

canvas.create_text(
    516.0,
    223.0,
    anchor="nw",
    text="Równanie Funkcji",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)
canvas.create_text(
    55.0,
    353.0,
    anchor="nw",
    text="Częstotliwość [Hz]",
    fill="#000000",
    font=("Inter Regular", 12 * -1)
)

button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: Break(),
    relief="flat"
)
button_1.place(
    x=499.0,
    y=324.0,
    width=181.0,
    height=46.0
)

button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: on_button2_clicked_Start(),
    relief="flat"
)
button_2.place(
    x=268.0,
    y=324.0,
    width=181.0,
    height=46.0
)

button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: calculateCx(),
    relief="flat"
)
button_3.place(
    x=509.0,
    y=96.0,
    width=181.0,
    height=46.0
)

button_image_4 = PhotoImage(
    file=relative_to_assets("button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: Export(),
    relief="flat"
)
button_4.place(
    x=382.0,
    y=407.0,
    width=186.0,
    height=42.0
)

window.resizable(False, False)
window.mainloop()